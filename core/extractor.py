import concurrent.futures
import io
import os
import time

import openpyxl
import pdfplumber
import streamlit as st

from config import settings

# ─── SparkSession singleton ───────────────────────────────────────────────────

_spark = None

# Palavras-chave que indicam erros transientes de cluster (inicializando / indisponível)
_ERROS_TRANSIENTES = ("UNAVAILABLE", "TIMEOUT", "DEADLINE", "CONNECTION", "CLUSTER",
                      "UNREACHABLE", "REFUSED", "RESET")

# Registro das últimas queries executadas (reset a cada chamada não-cacheada)
_recent_sql: list[str] = []


def get_recent_sql() -> list[str]:
    """Retorna as queries SQL executadas desde o último clear_recent_sql()."""
    return list(_recent_sql)


def clear_recent_sql() -> None:
    """Limpa o registro de queries recentes."""
    _recent_sql.clear()


def _get_spark(timeout_s: int = 10):
    global _spark
    if _spark is not None:
        return _spark
    from databricks.connect import DatabricksSession

    def _connect():
        return DatabricksSession.builder.remote(
            host=settings.databricks_host,
            token=settings.databricks_token,
            cluster_id=settings.databricks_cluster_id,
        ).getOrCreate()

    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
        fut = ex.submit(_connect)
        try:
            _spark = fut.result(timeout=timeout_s)
        except concurrent.futures.TimeoutError:
            raise TimeoutError(
                f"Cluster Databricks não respondeu em {timeout_s}s. "
                "O cluster pode estar iniciando — aguarde e tente novamente."
            )
    return _spark


_ERROS_SESSAO_OBSOLETA = ("SESSION_CHANGED", "NO_ACTIVE_SESSION", "INVALID_SESSION", "SESSION_NOT_FOUND")

def _spark_sql(query: str):
    """Executa apenas SELECT no Spark. Reconecta automaticamente se a sessão ficar obsoleta."""
    normalized = query.strip().lstrip("(").lstrip().upper()
    if not normalized.startswith(("SELECT", "WITH", "SHOW", "DESCRIBE")):
        raise ValueError(f"Apenas consultas SELECT são permitidas. Query bloqueada: {query[:80]}")
    global _spark
    _recent_sql.append(query)
    for attempt in range(2):
        try:
            return _get_spark().sql(query)
        except Exception as e:
            msg = str(e).upper()
            if attempt == 0 and any(k in msg for k in _ERROS_SESSAO_OBSOLETA):
                _spark = None
                continue
            raise


def aguardar_cluster(max_wait: int = 300, step: int = 30) -> bool:
    """Tenta conectar ao cluster em loop até max_wait segundos. Retorna True se conectado."""
    global _spark
    deadline = time.time() + max_wait
    while time.time() < deadline:
        try:
            _spark = None  # força re-conexão a cada tentativa
            _get_spark().sql("SELECT 1").collect()
            return True
        except Exception as e:
            msg = str(e).upper()
            if not any(k in msg for k in _ERROS_TRANSIENTES):
                raise  # erro de config/permissão — não tem retry
            time.sleep(step)
    return False


def _is_erro_transiente(e: Exception) -> bool:
    if isinstance(e, (TimeoutError, concurrent.futures.TimeoutError)):
        return True
    msg = str(e).upper()
    return any(k in msg for k in _ERROS_TRANSIENTES)


def _databricks_configurado() -> bool:
    return bool(settings.databricks_host and settings.databricks_token and settings.databricks_cluster_id)


# ─── Funções Databricks ───────────────────────────────────────────────────────

def descrever_colunas_balancete() -> list[str]:
    """Retorna os nomes reais das colunas da tabela para diagnóstico."""
    cat = settings.databricks_catalog
    sch = settings.databricks_schema
    rows = _spark_sql(f"DESCRIBE {cat}.{sch}.balancete_consolidado").collect()
    return [row[0] for row in rows if row[0] and not row[0].startswith("#")]


@st.cache_data(ttl=3600)
def listar_referencias() -> list[int]:
    cat = settings.databricks_catalog
    sch = settings.databricks_schema
    df = _spark_sql(f"""
        SELECT Referencia FROM (
            SELECT DISTINCT Referencia FROM {cat}.{sch}.portal_sindico_contacorrente
            INTERSECT
            SELECT DISTINCT Referencia FROM {cat}.{sch}.balancete_consolidado
        )
        ORDER BY Referencia
    """)
    return [row[0] for row in df.collect()]


@st.cache_data(ttl=1800)
def extrair_de_databricks(
    referencia: int,
    n_meses: int = 2,
    periodos_fixos: tuple[str, ...] | None = None,
) -> str:
    cat = settings.databricks_catalog
    sch = settings.databricks_schema

    if periodos_fixos:
        # Valida quais períodos existem de fato na tabela para o condomínio
        lista = ", ".join(f"'{p}'" for p in periodos_fixos)
        periodos_df = _spark_sql(f"""
            SELECT DISTINCT Mes_Referencia
            FROM {cat}.{sch}.balancete_consolidado
            WHERE Referencia = {referencia}
              AND Mes_Referencia IN ({lista})
            ORDER BY
                CAST(split(Mes_Referencia, '/')[1] AS INT) ASC,
                CAST(split(Mes_Referencia, '/')[0] AS INT) ASC
        """).collect()
        periodos = [row[0] for row in periodos_df]
    else:
        # N períodos mais recentes do balancete
        # Mes_Referencia está em formato MM/yyyy — ordena por ano e mês para garantir ordem cronológica
        periodos_df = _spark_sql(f"""
            SELECT DISTINCT Mes_Referencia
            FROM {cat}.{sch}.balancete_consolidado
            WHERE Referencia = {referencia}
            ORDER BY
                CAST(split(Mes_Referencia, '/')[1] AS INT) DESC,
                CAST(split(Mes_Referencia, '/')[0] AS INT) DESC
            LIMIT {n_meses}
        """).collect()
        periodos = [row[0] for row in periodos_df]

    if not periodos:
        return ""

    partes = []

    for periodo in reversed(periodos):
        # Passo 2: Itens do balancete para o período
        bal_rows = _spark_sql(f"""
            SELECT Categoria_Balancete, Descricao, Valor,
                   Saldo_Anterior, Credito, Debito, Saldo_Final
            FROM {cat}.{sch}.balancete_consolidado
            WHERE Referencia = {referencia}
              AND Mes_Referencia = '{periodo}'
            ORDER BY Categoria_Balancete, Descricao
        """).collect()

        if bal_rows:
            partes.append(f"=== BALANCETE CONSOLIDADO - {periodo} ===")
            partes.append("Categoria | Descrição | Valor | Saldo_Anterior | Crédito | Débito | Saldo_Final")
            for r in bal_rows:
                partes.append(
                    f"{r['Categoria_Balancete']} | {r['Descricao']} | {r['Valor']} | "
                    f"{r['Saldo_Anterior']} | {r['Credito']} | {r['Debito']} | {r['Saldo_Final']}"
                )

        # Passo 3: Conta corrente do mesmo mês
        # Mes_Referencia está em formato MM/yyyy (ex: "11/2025")
        cc_rows = _spark_sql(f"""
            SELECT Conta_Principal, Historico, Debito, Credito, Data
            FROM {cat}.{sch}.portal_sindico_contacorrente
            WHERE Referencia = {referencia}
              AND month(Data) = month(to_date('{periodo}', 'MM/yyyy'))
              AND year(Data) = year(to_date('{periodo}', 'MM/yyyy'))
            ORDER BY Data, Conta_Principal
        """).collect()

        if cc_rows:
            partes.append(f"=== CONTA CORRENTE - {periodo} ===")
            partes.append("Conta Principal | Histórico | Débito | Crédito | Data")
            for r in cc_rows:
                partes.append(
                    f"{r['Conta_Principal']} | {r['Historico']} | "
                    f"{r['Debito']} | {r['Credito']} | {r['Data']}"
                )

    return "\n".join(partes).strip()


def extrair_texto(source) -> str:
    """Dispatcher: detecta o tipo pelo nome/extensão e chama o extrator correto."""
    nome = source if isinstance(source, (str, os.PathLike)) else getattr(source, "name", "")
    if str(nome).lower().endswith(".xlsx"):
        return extrair_texto_xlsx(source)
    return extrair_texto_pdf(source)


def extrair_texto_pdf(source) -> str:
    """Aceita caminho (str/Path) ou objeto de upload do Streamlit."""
    if isinstance(source, (str, os.PathLike)):
        ctx = pdfplumber.open(source)
    else:
        ctx = pdfplumber.open(io.BytesIO(source.read()))
    with ctx as pdf:
        return "\n".join(
            page.extract_text() or "" for page in pdf.pages
        ).strip()


def extrair_texto_xlsx(source) -> str:
    """Lê todas as abas do XLSX e retorna o conteúdo como texto estruturado."""
    if isinstance(source, (str, os.PathLike)):
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(source.read()), read_only=True, data_only=True)

    partes = []
    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        partes.append(f"=== Planilha: {nome_aba} ===")
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                partes.append(" | ".join("" if cell is None else str(cell) for cell in row))
    wb.close()
    return "\n".join(partes).strip()
